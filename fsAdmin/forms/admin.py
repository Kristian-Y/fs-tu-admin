from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
import requests
from .models import SponsorApplication, MemberApplication, ContactMessage



@admin.register(SponsorApplication)
class SponsorApplicationAdmin(admin.ModelAdmin):
    list_display = ("company_name", "contact_person_names", "email", "tier", "created_at")
    search_fields = ("company_name", "contact_person_names", "email")
    list_filter = ("tier", "created_at")
    actions = ["export_as_xlsx"]


    def export_as_xlsx(self, request, queryset):
        """Експортира избраните записи в XLSX"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sponsor Applications"

        # Заглавия
        headers = ["Company Name", "Contact Person", "Email", "Phone", "Tier", "Description", "Created At"]
        ws.append(headers)

        # Данни
        for app in queryset:
            ws.append([
                app.company_name,
                app.contact_person_names,
                app.email,
                app.phone,
                app.tier,
                app.description,
                app.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        # Визуално форматиране (по желание)
        for col in ws.columns:
            max_length = 0
            col_name = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_name].width = max_length + 2

        # Отговор
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="sponsor_applications.xlsx"'
        wb.save(response)
        return response

    export_as_xlsx.short_description = "📊 Export as XLSX"


@admin.register(MemberApplication)
class MemberApplicationAdmin(admin.ModelAdmin):
    list_display = ("first_name", "last_name", "email", "university", "major", "created_at")
    search_fields = ("first_name", "last_name", "email", "university", "major")
    list_filter = ("university", "created_at")
    actions = ["export_as_xlsx"]

    def export_as_xlsx(self, request, queryset):
        """Експортира избраните записи в XLSX"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Member Applications"

        headers = ["First Name", "Last Name", "Email", "Phone", "University", "Major", "Graduation",
                   "Course", "Semester", "Skills", "Motivation", "Portfolio Link", "Created At"]
        ws.append(headers)

        for app in queryset:
            ws.append([
                app.first_name,
                app.last_name,
                app.email,
                app.phone,
                app.university,
                app.major,
                app.graduation,
                app.course,
                app.semester,
                app.skills,
                app.motivation,
                app.portfolio_link,
                app.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        for col in ws.columns:
            max_length = 0
            col_name = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_name].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="member_applications.xlsx"'
        wb.save(response)
        return response

    export_as_xlsx.short_description = "📊 Export as XLSX"

@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ("name", "email", "subject", "created_at")
    search_fields = ("name", "email", "subject")
    list_filter = ("created_at",)
    actions = ["export_as_xlsx"]

    def export_as_xlsx(self, request, queryset):
        """Експортира избраните съобщения в XLSX"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Contact Messages"

        headers = ["Name", "Email", "Subject", "Message", "Created At"]
        ws.append(headers)

        for msg in queryset:
            ws.append([
                msg.name,
                msg.email,
                msg.subject,
                msg.message,
                msg.created_at.strftime("%Y-%m-%d %H:%M"),
            ])

        for col in ws.columns:
            max_length = 0
            col_name = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_name].width = max_length + 2

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="contact_messages.xlsx"'
        wb.save(response)
        return response

    export_as_xlsx.short_description = "📊 Export as XLSX"